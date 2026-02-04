import streamlit as st
import pandas as pd
import os
import io
import zipfile
import tempfile
import sys
import xlrd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# ===== PAGE CONFIG =====
st.set_page_config(
    page_title="Excel Merger & Report Generator",
    page_icon="📊",
    layout="wide"
)

# ===== STYLES =====
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 700;
        color: #1E3A5F;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1.1rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }
    .step-header {
        font-size: 1.3rem;
        font-weight: 600;
        color: #2E7D32;
        margin-top: 1.5rem;
    }
    .success-box {
        padding: 1rem;
        background-color: #E8F5E9;
        border-radius: 0.5rem;
        border-left: 4px solid #4CAF50;
    }
    .info-box {
        padding: 1rem;
        background-color: #E3F2FD;
        border-radius: 0.5rem;
        border-left: 4px solid #2196F3;
    }
</style>
""", unsafe_allow_html=True)

# ===== CONSTANTS =====
MONTH_ORDER = [
    "mar 25", "mar25", "MAR 25", "apr 25", "apr25", "APR 25",
    "may 25", "may25", "MAY 25", "jun 25", "jun25", "JUN 25",
    "jul 25", "jul25", "JUL 25", "aug 25", "aug25", "AUG 25",
    "sep 25", "sep25", "SEP 25", "oct 25", "oct25", "OCT 25",
    "nov 25", "nov25", "NOV 25", "dec 25", "dec25", "DEC 25",
    "jan 26", "jan26", "JAN 26", "feb 26", "feb26", "FEB 26"
]

MONTH_DISPLAY = {
    "mar 25": "Mar-2025", "mar25": "Mar-2025", "MAR 25": "Mar-2025",
    "apr 25": "Apr-2025", "apr25": "Apr-2025", "APR 25": "Apr-2025",
    "may 25": "May-2025", "may25": "May-2025", "MAY 25": "May-2025",
    "jun 25": "Jun-2025", "jun25": "Jun-2025", "JUN 25": "Jun-2025",
    "jul 25": "Jul-2025", "jul25": "Jul-2025", "JUL 25": "Jul-2025",
    "aug 25": "Aug-2025", "aug25": "Aug-2025", "AUG 25": "Aug-2025",
    "sep 25": "Sep-2025", "sep25": "Sep-2025", "SEP 25": "Sep-2025",
    "oct 25": "Oct-2025", "oct25": "Oct-2025", "OCT 25": "Oct-2025",
    "nov 25": "Nov-2025", "nov25": "Nov-2025", "NOV 25": "Nov-2025",
    "dec 25": "Dec-2025", "dec25": "Dec-2025", "DEC 25": "Dec-2025",
    "jan 26": "Jan-2026", "jan26": "Jan-2026", "JAN 26": "Jan-2026",
    "feb 26": "Feb-2026", "feb26": "Feb-2026", "FEB 26": "Feb-2026"
}

EXCLUDE_COLUMNS = [
    "Unnamed: 0", "SR.NO", "Source_File", "Month", "Month_Order",
    "EMPLOYEE NAME", "GENDER M/F", "NAME OF SCHOOL",
    "GROSS AFTER DEDUCTING FA", "GROSS PAYMENT AFTER GOVT DEDUCTIONS",
    "GROSS PAYMENT AFTER NPS DEDUCTIONS", "NGR(TOTAL DEDUCTIONS)",
    "EMPLOYEE NET SALARY", "TOTAL GOVT DEDUCTIONS", "NPS TOTAL"
]

INFO_COLUMNS = [
    "BLOCK / TALUKA", "SCHOOL UDISE CODE", "SCHOOL SHALARTH DDO CODE",
    "S.R NO OF EMPL", "SHALARTH ID", "DESIGNATION ", "GPF NO", "DCPS NO",
    "PRAN NO", "PAN NO", "ADHAR NO", "MOB NO", "EMAIL ID",
    "DDO BANK NAME", "DDO BANK ACCOUNT NUMBER", "DDO BANK IFSC CODE",
    "BANK NAME", "BANK ACCOUNT NUMBER", "BANK IFSC CODE", "BRANCH NAME",
    "PAY MATRIX", "REMARKS"
]

NUMERIC_COLUMNS = [
    "BASIC PAY", "D.A", "HRA", "T.A", "T.A ARREARS", "TRIBAL ALLOWANCE",
    "WASHING ALLOWANCE ", "DA ARREARS ", "HRA ARREARS ", "BASIC ARREARS ",
    "CLA", "NPS EMPR ALLOW", "TOTAL PAY", "F A",
    "GPF", "GPF ADV", "PT", "GIS(ZP)", "GIS SCOUT", "DCPS REGULAR",
    "DCPS DELAYED", "DCPS PAY ARREARS RECOVERY", "REVENUE STAMP",
    "DCPS DA ARREARS RECOVERY", "GROUP ACCIDENTAL POLICY", "NAA",
    "TOTAL GOVT DEDUCTIONS",
    "NPS EMPR CONTRI", "NPS EMP CONTRI", "NPS EMPR CONTRI ARR",
    "NPS EMP CONTRI ARR", "NPS TOTAL",
    "INCOME TAX", "CO-OP BANK", "NGR(LIC)", "NGR(SOCIETY LOAN)", "NGR(MISC)",
    "NGR(OTHER RECOVERY)", "NGR(RD)", "NGR(OTHER DEDUCTION)"
]


# ===== HELPER FUNCTIONS =====

class VirtualUploadedFile:
    """Minimal file-like wrapper for zip entries."""
    def __init__(self, name, content):
        self.name = name
        self._buffer = io.BytesIO(content)

    def read(self):
        return self._buffer.read()

    def seek(self, pos):
        self._buffer.seek(pos)


def extract_archive_files(archive_file):
    """Extract .xls files from a zip upload and return file-like objects."""
    extracted = []
    name = getattr(archive_file, "name", "").lower()
    content = archive_file.read()
    archive_file.seek(0)

    if name.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                if info.filename.lower().endswith(".xls"):
                    content = zf.read(info.filename)
                    extracted.append(VirtualUploadedFile(info.filename, content))
        return extracted

    st.error("Unsupported archive type. Please upload a .zip file.")
    return []


def save_uploaded_files(selected_files, root_dir, school_name):
    """Save uploaded files into month folders under root_dir/school_name."""
    base_path = os.path.join(root_dir, school_name)
    os.makedirs(base_path, exist_ok=True)

    for uploaded_file in selected_files:
        parts = uploaded_file.name.replace("\\", "/").split("/")
        if len(parts) >= 2:
            month = parts[-2]
        else:
            month = "Unknown"

        month_path = os.path.join(base_path, month)
        os.makedirs(month_path, exist_ok=True)

        file_name = parts[-1]
        file_path = os.path.join(month_path, file_name)

        content = uploaded_file.read()
        uploaded_file.seek(0)

        with open(file_path, "wb") as f:
            f.write(content)

    return base_path


def run_scripts(root_dir, school_name):
    """Run merge_alternate.py then report.py using temp root directory."""
    env = os.environ.copy()
    env["EXCEL_MERGER_ROOT"] = root_dir

    merge_cmd = [sys.executable, "merge_alternate.py", school_name]
    report_cmd = [sys.executable, "report.py", school_name]

    merge_result = subprocess.run(
        merge_cmd,
        capture_output=True,
        text=True,
        env=env
    )
    if merge_result.returncode != 0:
        raise RuntimeError(merge_result.stderr or merge_result.stdout)

    report_result = subprocess.run(
        report_cmd,
        capture_output=True,
        text=True,
        env=env
    )
    if report_result.returncode != 0:
        raise RuntimeError(report_result.stderr or report_result.stdout)


def zip_folder(folder_path):
    """Create a ZIP of a folder and return bytes."""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for root, _, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, folder_path)
                zip_file.write(file_path, arcname=arcname)
    zip_buffer.seek(0)
    return zip_buffer


def find_header_row(file_content):
    """Find the header row containing SR.NO"""
    preview = pd.read_excel(io.BytesIO(file_content), engine="xlrd", header=None, nrows=40)
    for i, row in preview.iterrows():
        if row.astype(str).str.upper().isin(["SR.NO"]).any():
            return i
    raise ValueError("SR.NO not found in file")


def process_uploaded_files(uploaded_files, school_name, progress_bar, status_text):
    """Process uploaded files and return merged data by month"""
    
    # Group files by month folder
    month_files = {}
    for uploaded_file in uploaded_files:
        # Extract month from path (e.g., "APR 25/file.xls" -> "APR 25")
        parts = uploaded_file.name.replace("\\", "/").split("/")
        if len(parts) >= 2:
            month = parts[-2]
        else:
            month = "Unknown"
        
        if month not in month_files:
            month_files[month] = []
        month_files[month].append(uploaded_file)
    
    all_data = {}
    total_months = len(month_files)
    
    for idx, (month, files) in enumerate(sorted(month_files.items())):
        status_text.text(f"Processing {month}...")
        progress_bar.progress((idx + 1) / total_months)
        
        month_tables = []
        for file in files:
            try:
                content = file.read()
                file.seek(0)  # Reset for potential re-read
                
                header_row = find_header_row(content)
                df = pd.read_excel(io.BytesIO(content), engine="xlrd", header=header_row)
                df = df.dropna(how="all")
                
                # Remove Grand Total rows
                if "SR.NO" in df.columns:
                    df = df[df["SR.NO"].astype(str).str.upper() != "GRAND TOTAL"]
                
                df["Source_File"] = file.name.split("/")[-1]
                df["Month"] = month
                month_tables.append(df)
            except Exception as e:
                st.warning(f"Error processing {file.name}: {e}")
        
        if month_tables:
            all_data[month] = pd.concat(month_tables, ignore_index=True)
    
    return all_data


def create_merged_workbook(all_data, school_name):
    """Create merged Excel workbook with monthly sheets"""
    wb = Workbook()
    wb.remove(wb.active)
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for month, df in sorted(all_data.items()):
        ws = wb.create_sheet(title=month[:31])  # Excel sheet name limit
        
        # Write header
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Write data
        for row_idx, row in enumerate(df.values, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Auto-width columns
        for col_idx in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    return wb


def get_value(row, col_name):
    """Safely get value from row"""
    try:
        val = row.get(col_name, 0)
        if pd.isna(val) or val == "" or val is None:
            return 0
        return val
    except:
        return 0


def is_zero_or_empty(val):
    """Check if value is zero or empty"""
    if val is None or val == "" or pd.isna(val):
        return True
    try:
        return float(val) == 0
    except:
        return False


def get_non_zero_columns(emp_df, numeric_cols):
    """Get columns with at least one non-zero value"""
    non_zero_cols = []
    for col in numeric_cols:
        if col in emp_df.columns:
            has_non_zero = any(not is_zero_or_empty(val) for val in emp_df[col])
            if has_non_zero:
                non_zero_cols.append(col)
    return non_zero_cols


def create_february_row(jan_row, active_columns):
    """Create February row based on January data"""
    feb_row = jan_row.copy()
    feb_row["Month"] = "FEB 26"
    
    pt_val = get_value(jan_row, "PT")
    if pt_val == 200:
        feb_row["PT"] = 300
    elif pt_val == 0 or pt_val == "" or pd.isna(pt_val):
        feb_row["PT"] = 0
    else:
        feb_row["PT"] = 300
    
    feb_row["GROUP ACCIDENTAL POLICY"] = 531
    if "INCOME TAX" in feb_row.index:
        feb_row["INCOME TAX"] = 0
    return feb_row


def create_employee_report(emp_name, emp_df, all_numeric_cols):
    """Create individual employee report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Tax Details"
    
    jan_variants = ["jan26", "jan 26", "JAN 26"]
    feb_variants = ["feb26", "feb 26", "FEB 26"]

    # Sort by month
    emp_df = emp_df.copy()
    emp_df["Month_Order"] = emp_df["Month"].apply(
        lambda x: MONTH_ORDER.index(x) if x in MONTH_ORDER else 99
    )
    emp_df = emp_df.sort_values("Month_Order")

    if "INCOME TAX" in emp_df.columns and "Month" in emp_df.columns:
        emp_df.loc[emp_df["Month"].isin(feb_variants), "INCOME TAX"] = 0
    
    first_row = emp_df.iloc[0]
    
    # Check for January/February
    has_jan = any(v in emp_df["Month"].values for v in jan_variants)
    has_feb = any(v in emp_df["Month"].values for v in feb_variants)
    
    jan_month = None
    for v in jan_variants:
        if v in emp_df["Month"].values:
            jan_month = v
            break
    
    # Get non-zero columns
    active_numeric_cols = get_non_zero_columns(emp_df, all_numeric_cols)

    # Always keep INCOME TAX column even if all values are zero
    if "INCOME TAX" in all_numeric_cols and "INCOME TAX" in emp_df.columns:
        if "INCOME TAX" not in active_numeric_cols:
            active_numeric_cols.append("INCOME TAX")
    
    if has_jan and not has_feb and jan_month:
        jan_data = emp_df[emp_df["Month"] == jan_month].iloc[0]
        pt_val = get_value(jan_data, "PT")
        if pt_val == 200 or (pt_val and pt_val != 0):
            if "PT" not in active_numeric_cols:
                active_numeric_cols.append("PT")
        if "GROUP ACCIDENTAL POLICY" not in active_numeric_cols:
            active_numeric_cols.append("GROUP ACCIDENTAL POLICY")
    
    active_numeric_cols = [col for col in all_numeric_cols if col in active_numeric_cols]
    
    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=9)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Row 1: Title
    total_cols = 2 + len(active_numeric_cols)
    ws.cell(row=1, column=1, value="INCOME TAX DETAILS 2024-25")
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = center_align
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    
    # Row 2: Employee Name + School
    gender = str(first_row.get("GENDER M/F", "")).strip().upper()
    salutation = "SHRI" if gender == "M" else "SHRIMATI"
    
    ws.cell(row=2, column=1, value=f"{salutation} {emp_name}")
    ws.cell(row=2, column=1).font = Font(bold=True)
    
    school_name = first_row.get("NAME OF SCHOOL", "")
    if school_name:
        school_col = max(total_cols - 4, 1)
        ws.cell(row=2, column=school_col, value=str(school_name))
    
    # Row 3: Headers
    headers = ["SR.NO", "MONTH"] + active_numeric_cols
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
    
    # Data rows
    current_row = 4
    sr_no = 1
    
    for _, row in emp_df.iterrows():
        month_name = MONTH_DISPLAY.get(row["Month"], row["Month"])
        data = [sr_no, month_name]
        for col in active_numeric_cols:
            if col == "INCOME TAX" and row.get("Month") in feb_variants:
                data.append(0)
            else:
                data.append(get_value(row, col))
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal='right')
        
        sr_no += 1
        current_row += 1
    
    # February row
    if has_jan and not has_feb and jan_month:
        jan_data = emp_df[emp_df["Month"] == jan_month].iloc[0]
        feb_data = create_february_row(jan_data, active_numeric_cols)
        
        data = [sr_no, "Feb-2026"]
        for col in active_numeric_cols:
            if col == "PT":
                data.append(feb_data["PT"])
            elif col == "GROUP ACCIDENTAL POLICY":
                data.append(531)
            elif col == "INCOME TAX":
                data.append(0)
            else:
                data.append(get_value(feb_data, col))
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal='right')
        
        sr_no += 1
        current_row += 1
    
    # Total row
    def safe_sum(col_name):
        try:
            ser = emp_df[col_name].fillna(0)
            if col_name == "INCOME TAX" and "Month" in emp_df.columns:
                ser = ser.where(~emp_df["Month"].isin(feb_variants), 0)
            total = ser.sum()
            if has_jan and not has_feb and jan_month:
                jan_data = emp_df[emp_df["Month"] == jan_month].iloc[0]
                feb_data = create_february_row(jan_data, active_numeric_cols)
                if col_name == "PT":
                    total += feb_data["PT"]
                elif col_name == "GROUP ACCIDENTAL POLICY":
                    total += 531
                elif col_name == "INCOME TAX":
                    total += 0
                else:
                    feb_val = get_value(feb_data, col_name)
                    if feb_val:
                        try:
                            total += float(feb_val)
                        except:
                            pass
            return total
        except:
            return 0
    
    totals = ["", "Total"]
    for col in active_numeric_cols:
        totals.append(safe_sum(col))
    
    for col_idx, value in enumerate(totals, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=value)
        cell.border = thin_border
        cell.font = Font(bold=True)
    
    current_row += 2
    
    # Footer
    ws.cell(row=current_row, column=1, 
            value="या तक्त्यात काही चूक आढळून आल्यास तात्काळ मुख्याध्यापकांच्या लक्षात आणून द्यावी, नजरचुकीने काही चूक झाल्यास लागणा-या आयकरास कर्मचारी स्वत: जबाबदार राहील.")
    current_row += 2
    
    ws.cell(row=current_row, column=1, value="Employee Signature")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    
    headmaster_col = max(14, total_cols - 3)
    ws.cell(row=current_row, column=headmaster_col, value="Headmaster")
    ws.cell(row=current_row, column=headmaster_col).font = Font(bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=1, value=f"{salutation} {emp_name}")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    
    # Page setup
    TOTAL_PAGE_WIDTH = 140
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 11
    
    remaining_cols = total_cols - 2
    if remaining_cols > 0:
        remaining_width = TOTAL_PAGE_WIDTH - 6 - 11
        col_width = max(8, min(15, remaining_width / remaining_cols))
        for i in range(3, total_cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = col_width
    
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)
    
    last_col = get_column_letter(total_cols)
    ws.print_area = f'A1:{last_col}{current_row}'
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    
    return wb


def generate_all_reports(all_data, progress_bar, status_text):
    """Generate reports for all employees"""
    # Combine all months
    combined_df = pd.concat(all_data.values(), ignore_index=True)
    
    # Get numeric columns
    actual_numeric_cols = [col for col in NUMERIC_COLUMNS if col in combined_df.columns and col not in EXCLUDE_COLUMNS]
    
    # Find employee name column
    EMP_NAME_COL = "EMPLOYEE NAME"
    combined_df[EMP_NAME_COL] = combined_df[EMP_NAME_COL].astype(str).str.strip().str.upper()
    combined_df = combined_df[combined_df[EMP_NAME_COL] != ""]
    combined_df = combined_df[combined_df[EMP_NAME_COL] != "NAN"]
    combined_df = combined_df[~combined_df[EMP_NAME_COL].str.contains("GRAND TOTAL", case=False, na=False)]
    
    employees = combined_df[EMP_NAME_COL].unique()
    reports = {}
    
    for idx, emp_name in enumerate(employees):
        if not emp_name or emp_name == "NAN":
            continue
        
        status_text.text(f"Generating report for {emp_name}...")
        progress_bar.progress((idx + 1) / len(employees))
        
        try:
            emp_df = combined_df[combined_df[EMP_NAME_COL] == emp_name]
            wb = create_employee_report(emp_name, emp_df, actual_numeric_cols)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Safe filename
            safe_name = emp_name.replace("/", "_").replace("\\", "_").replace(":", "_")
            safe_name = safe_name.replace("*", "_").replace("?", "_").replace('"', "_")
            safe_name = safe_name.replace("<", "_").replace(">", "_").replace("|", "_")
            
            reports[f"{safe_name}.xlsx"] = output.getvalue()
        except Exception as e:
            st.warning(f"Error creating report for {emp_name}: {e}")
    
    return reports


def create_zip(files_dict):
    """Create a ZIP file from dictionary of files"""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for filename, content in files_dict.items():
            zip_file.writestr(filename, content)
    zip_buffer.seek(0)
    return zip_buffer


# ===== MAIN APP =====

st.markdown('<p class="main-header">📊 Excel Merger & Report Generator</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">Upload school data, merge Excel files, and generate individual employee income tax reports</p>', unsafe_allow_html=True)

# School name input
school_name = st.text_input("🏫 School Name", placeholder="e.g., PALASGAON or KANYA BASMATH")

st.markdown("---")

# File upload section
st.markdown('<p class="step-header">Step 1: Upload Excel Files</p>', unsafe_allow_html=True)

st.info("""
**How to upload:**
1. Select all .xls files from your school folder (including all month subfolders)
2. Make sure each file path includes the month folder name (e.g., `APR 25/file.xls`)
3. Or upload a single .zip of the whole school folder

**Expected folder structure:**
```
School Folder/
├── APR 25/
│   ├── file1.xls
│   └── file2.xls
├── MAY 25/
│   ├── file1.xls
│   └── file2.xls
...
```
""")

zip_upload = st.file_uploader(
    "Upload school folder (.zip)",
    type=["zip"],
    accept_multiple_files=False,
    help="Zip the school folder with month subfolders, then upload"
)

uploaded_files = st.file_uploader(
    "Upload Excel files (.xls)",
    type=["xls"],
    accept_multiple_files=True,
    help="Select all Excel files from all month folders"
)

selected_files = None
if zip_upload is not None:
    selected_files = extract_archive_files(zip_upload)
elif uploaded_files:
    selected_files = uploaded_files

if selected_files and school_name:
    st.success(f"Uploaded {len(selected_files)} files for **{school_name}**")
    
    # Show file preview
    with st.expander("📁 View uploaded files"):
        for f in selected_files[:20]:
            st.text(f"• {f.name}")
        if len(selected_files) > 20:
            st.text(f"... and {len(selected_files) - 20} more files")
    
    st.markdown("---")
    
    # Process button
    if st.button("🚀 Process Files", type="primary", use_container_width=True):
        
        st.markdown('<p class="step-header">Step 2: Preparing Files...</p>', unsafe_allow_html=True)
        status_text = st.empty()
        status_text.text("Saving uploaded files to a temporary folder...")
        
        try:
            with tempfile.TemporaryDirectory() as temp_root:
                save_uploaded_files(selected_files, temp_root, school_name)
                
                status_text.text("Running merge and report scripts...")
                run_scripts(temp_root, school_name)
                
                merged_path = os.path.join(temp_root, f"{school_name}_Merged_Monthly.xlsx")
                reports_dir = os.path.join(temp_root, f"{school_name}_income_tax_reports")
                
                if not os.path.exists(merged_path):
                    st.error("Merged file not found. Please check your inputs.")
                    st.stop()
                
                status_text.text("Preparing downloads...")
                with open(merged_path, "rb") as f:
                    merged_bytes = f.read()
                
                if not os.path.isdir(reports_dir):
                    st.error("Reports folder not found. Please check your inputs.")
                    st.stop()
                
                reports_zip = zip_folder(reports_dir)
                
                st.markdown(f"""
                <div class="success-box">
                    <strong>✅ Process Complete!</strong><br>
                    • Merged file and reports are ready
                </div>
                """, unsafe_allow_html=True)
                
                st.download_button(
                    label="📥 Download Merged Excel",
                    data=merged_bytes,
                    file_name=f"{school_name}_Merged_Monthly.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.download_button(
                    label="📥 Download All Reports (ZIP)",
                    data=reports_zip.getvalue(),
                    file_name=f"{school_name}_Income_Tax_Reports.zip",
                    mime="application/zip"
                )
                
                st.balloons()
        except Exception as e:
            st.error(f"Processing failed: {e}")

elif selected_files and not school_name:
    st.warning("Please enter the school name to continue.")

# Footer
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #888; font-size: 0.9rem;">
    Made for easy Excel merging and Income Tax report generation
</div>
""", unsafe_allow_html=True)
