import pandas as pd
import os
import sys
import xlrd
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# PDF generation using Excel automation
import win32com.client
from PyPDF2 import PdfMerger

# ===== CONFIGURATION =====
# Usage: python report.py [folder_name]
# Example: python report.py palasgaon
# Example: python report.py "KANYA BASMATH"

ROOT_DIR = os.environ.get("EXCEL_MERGER_ROOT", r"D:\excel merger")

# Get folder name from command line or use default
if len(sys.argv) > 1:
    SCHOOL_FOLDER = sys.argv[1]
else:
    SCHOOL_FOLDER = "palasgaon"  # Default

# Paths based on school
MERGED_FILE = os.path.join(ROOT_DIR, f"{SCHOOL_FOLDER}_Merged_Monthly.xlsx")
OUTPUT_DIR = os.path.join(ROOT_DIR, f"{SCHOOL_FOLDER}_income_tax_reports")
TEMPLATE_FILE = os.path.join(ROOT_DIR, "DESHMUKH SURYAKANT NARAYANRAO.xls")

print(f"School: {SCHOOL_FOLDER}")
print(f"Merged File: {MERGED_FILE}")
print(f"Output Dir: {OUTPUT_DIR}")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# Month order for sorting - handles both naming styles (apr25, APR 25)
MONTH_ORDER = [
    "mar 25", "mar25", "MAR 25",
    "apr 25", "apr25", "APR 25",
    "may 25", "may25", "MAY 25",
    "jun 25", "jun25", "JUN 25",
    "jul 25", "jul25", "JUL 25",
    "aug 25", "aug25", "AUG 25",
    "sep 25", "sep25", "SEP 25",
    "oct 25", "oct25", "OCT 25",
    "nov 25", "nov25", "NOV 25",
    "dec 25", "dec25", "DEC 25",
    "jan 26", "jan26", "JAN 26",
    "feb 26", "feb26", "FEB 26"
]

# Map folder names to display names (handles variations)
MONTH_DISPLAY = {
    "mar 25": "Mar-2025", "mar25": "Mar-2025", "MAR 25": "Mar-2025",
    "apr 25": "Apr-2025", "apr25": "Apr-2025", "APR 25": "Apr-2025",
    "may 25": "May-2025", "may25": "May-2025", "MAY 25": "May-2025",
    "jun 25": "Jun-2025", "jun25": "Jun-2025", "JUN 25": "Jun-2025",
    "jul 25": "Jul-2025", "jul25": "Jul-2025", "JUL 25": "Jul-2025",
    "aug 25": "Aug-2025", "aug25": "Aug-2025", "AUG 25": "Aug-2025",
    "sep 25": "Sep-2025", "sep25": "Sep-2025", "SEP 25": "Sep-2025",
    "oct 25": "Oct-2025", "oct25": "Oct-2025", "OCT 25": "Oct-2025",
    "nov 25": "Nov-2025", "nov25": "Nov-2025", "NOV 25": "Nov-2025",
    "dec 25": "Dec-2025", "dec25": "Dec-2025", "DEC 25": "Dec-2025",
    "jan 26": "Jan-2026", "jan26": "Jan-2026", "JAN 26": "Jan-2026",
    "feb 26": "Feb-2026", "feb26": "Feb-2026", "FEB 26": "Feb-2026"
}

# Columns to exclude from report (metadata columns, not data)
EXCLUDE_COLUMNS = [
    "Unnamed: 0", "SR.NO", "Source_File", "Month", "Month_Order",
    "EMPLOYEE NAME", "GENDER M/F", "NAME OF SCHOOL",
    # User requested to remove these columns from all reports:
    "GROSS AFTER DEDUCTING FA", "GROSS PAYMENT AFTER GOVT DEDUCTIONS",
    "GROSS PAYMENT AFTER NPS DEDUCTIONS", "NGR(TOTAL DEDUCTIONS)",
    "EMPLOYEE NET SALARY", "TOTAL GOVT DEDUCTIONS", "NPS TOTAL"
]

# Columns that are identifiers (show once, not per month)
INFO_COLUMNS = [
    "BLOCK / TALUKA", "SCHOOL UDISE CODE", "SCHOOL SHALARTH DDO CODE",
    "S.R NO OF EMPL", "SHALARTH ID", "DESIGNATION ", "GPF NO", "DCPS NO",
    "PRAN NO", "PAN NO", "ADHAR NO", "MOB NO", "EMAIL ID",
    "DDO BANK NAME", "DDO BANK ACCOUNT NUMBER", "DDO BANK IFSC CODE",
    "BANK NAME", "BANK ACCOUNT NUMBER", "BANK IFSC CODE", "BRANCH NAME",
    "PAY MATRIX", "REMARKS"
]

# Numeric columns (these will be shown per month and summed)
NUMERIC_COLUMNS = [
    "BASIC PAY", "D.A", "HRA", "T.A", "T.A ARREARS", "TRIBAL ALLOWANCE",
    "WASHING ALLOWANCE ", "DA ARREARS ", "HRA ARREARS ", "BASIC ARREARS ",
    "CLA", "NPS EMPR ALLOW", "TOTAL PAY", "F A",
    "GPF", "GPF ADV", "PT", "GIS(ZP)", "GIS SCOUT", "DCPS REGULAR",
    "DCPS DELAYED", "DCPS PAY ARREARS RECOVERY", "REVENUE STAMP",
    "DCPS DA ARREARS RECOVERY", "GROUP ACCIDENTAL POLICY", "NAA",
    "TOTAL GOVT DEDUCTIONS",
    "NPS EMPR CONTRI", "NPS EMP CONTRI", "NPS EMPR CONTRI ARR",
    "NPS EMP CONTRI ARR", "NPS TOTAL",
    "INCOME TAX", "CO-OP BANK", "NGR(LIC)", "NGR(SOCIETY LOAN)", "NGR(MISC)",
    "NGR(OTHER RECOVERY)", "NGR(RD)", "NGR(OTHER DEDUCTION)"
]


def find_header_row_xlsx(file_path, sheet_name=0):
    """Find the row containing SR.NO header in the merged xlsx file."""
    preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
    for i, row in preview.iterrows():
        if row.astype(str).str.upper().str.strip().isin(["SR.NO"]).any():
            return i
    return 0


def load_merged_data():
    """Load data from the merged Excel file (all sheets)."""
    print(f"Reading {MERGED_FILE}...")
    
    xlsx = pd.ExcelFile(MERGED_FILE)
    sheet_names = xlsx.sheet_names
    print(f"Found {len(sheet_names)} sheets: {sheet_names}")
    
    all_data = []
    
    for sheet_name in sheet_names:
        header_row = find_header_row_xlsx(MERGED_FILE, sheet_name)
        df = pd.read_excel(MERGED_FILE, sheet_name=sheet_name, header=header_row)
        df = df.dropna(how="all")
        
        if "Month" not in df.columns:
            df["Month"] = sheet_name
        
        all_data.append(df)
        print(f"  [{sheet_name}] - {len(df)} rows")
    
    combined_df = pd.concat(all_data, ignore_index=True)
    print(f"\nTotal records: {len(combined_df)}")
    
    return combined_df


def get_value(row_data, col_name):
    """Safely get value from row."""
    if col_name is None:
        return ""
    try:
        val = row_data.get(col_name, "")
        if pd.isna(val):
            return ""
        return val
    except:
        return ""


def is_zero_or_empty(val):
    """Check if value is 0, empty, or NaN."""
    if pd.isna(val):
        return True
    if val == "" or val == 0 or val == 0.0:
        return True
    try:
        if float(val) == 0:
            return True
    except:
        pass
    return False


def get_non_zero_columns(emp_df, numeric_cols):
    """Get list of numeric columns that have at least one non-zero value for this employee."""
    non_zero_cols = []
    for col in numeric_cols:
        if col in emp_df.columns:
            # Check if any value in this column is non-zero
            has_non_zero = False
            for val in emp_df[col]:
                if not is_zero_or_empty(val):
                    has_non_zero = True
                    break
            if has_non_zero:
                non_zero_cols.append(col)
    return non_zero_cols


def create_february_row(jan_row, active_columns):
    """Create February row based on January data with modifications."""
    feb_row = jan_row.copy()
    feb_row["Month"] = "feb26"
    
    # PT: 300 if other months have 200, else keep 0
    pt_val = get_value(jan_row, "PT")
    if pt_val == 200:
        feb_row["PT"] = 300
    elif pt_val == 0 or pt_val == "" or pd.isna(pt_val):
        feb_row["PT"] = 0
    else:
        feb_row["PT"] = 300  # Default to 300 if PT exists
    
    # ACC. INS. = 531 for Feb
    feb_row["GROUP ACCIDENTAL POLICY"] = 531

    # Income Tax should be 0 for February everywhere
    if "INCOME TAX" in feb_row.index:
        feb_row["INCOME TAX"] = 0
    
    return feb_row


def create_employee_report(emp_name, emp_df, all_numeric_cols):
    """Create individual employee report with all non-zero columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Tax Details"
    
    # Month variants used throughout the report
    jan_variants = ["jan26", "jan 26", "JAN 26"]
    feb_variants = ["feb26", "feb 26", "FEB 26"]

    # Sort by month
    emp_df = emp_df.copy()
    emp_df["Month_Order"] = emp_df["Month"].apply(
        lambda x: MONTH_ORDER.index(x) if x in MONTH_ORDER else 99
    )
    emp_df = emp_df.sort_values("Month_Order")

    # Income Tax must be 0 for February in all cases (even if source has values)
    if "INCOME TAX" in emp_df.columns and "Month" in emp_df.columns:
        emp_df.loc[emp_df["Month"].isin(feb_variants), "INCOME TAX"] = 0
    
    first_row = emp_df.iloc[0]
    
    # Check if January exists and February doesn't (handle different naming)
    has_jan = any(v in emp_df["Month"].values for v in jan_variants)
    has_feb = any(v in emp_df["Month"].values for v in feb_variants)
    
    # Find which January variant exists
    jan_month = None
    for v in jan_variants:
        if v in emp_df["Month"].values:
            jan_month = v
            break
    
    # Get columns that have non-zero values for this employee
    active_numeric_cols = get_non_zero_columns(emp_df, all_numeric_cols)

    # Always keep INCOME TAX column even if all values are zero
    if "INCOME TAX" in all_numeric_cols and "INCOME TAX" in emp_df.columns:
        if "INCOME TAX" not in active_numeric_cols:
            active_numeric_cols.append("INCOME TAX")

    # Always show Income Tax even if all values are zero
    if "INCOME TAX" in all_numeric_cols and "INCOME TAX" not in active_numeric_cols:
        active_numeric_cols.append("INCOME TAX")
    
    # If adding February, check if PT and GROUP ACCIDENTAL POLICY should be included
    if has_jan and not has_feb and jan_month:
        jan_data = emp_df[emp_df["Month"] == jan_month].iloc[0]
        pt_val = get_value(jan_data, "PT")
        if pt_val == 200 or (pt_val and pt_val != 0):
            if "PT" not in active_numeric_cols:
                active_numeric_cols.append("PT")
        if "GROUP ACCIDENTAL POLICY" not in active_numeric_cols:
            active_numeric_cols.append("GROUP ACCIDENTAL POLICY")
    
    # Preserve original order of columns
    active_numeric_cols = [col for col in all_numeric_cols if col in active_numeric_cols]
    
    # ===== STYLES =====
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=9)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ===== ROW 1: TITLE =====
    ws.cell(row=1, column=1, value="INCOME TAX DETAILS 2025-26")
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = center_align
    total_cols = 2 + len(active_numeric_cols)  # SR.NO + MONTH + numeric columns
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    
    # ===== ROW 2: EMPLOYEE NAME + SCHOOL =====
    gender = str(first_row.get("GENDER M/F", "")).strip().upper()
    salutation = "SHRI" if gender == "M" else "SHRIMATI"
    
    ws.cell(row=2, column=1, value=f"{salutation} {emp_name}")
    ws.cell(row=2, column=1).font = Font(bold=True)
    
    school_name = get_value(first_row, "NAME OF SCHOOL")
    school_col = min(10, total_cols)
    ws.cell(row=2, column=school_col, value=school_name)
    ws.cell(row=2, column=school_col).font = Font(bold=True)
    
    # ===== ROW 3: COLUMN HEADERS =====
    headers = ["SR. NO.", "MONTH"] + active_numeric_cols
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
    
    # ===== DATA ROWS =====
    current_row = 4
    sr_no = 1
    
    for _, row_data in emp_df.iterrows():
        month_display = MONTH_DISPLAY.get(row_data["Month"], row_data["Month"])
        
        data = [sr_no, month_display]
        for col in active_numeric_cols:
            if col == "INCOME TAX" and row_data.get("Month") in feb_variants:
                data.append(0)
            else:
                data.append(get_value(row_data, col))
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal='right')
        
        sr_no += 1
        current_row += 1
    
    # ===== ADD FEBRUARY ROW (if January exists but February doesn't) =====
    if has_jan and not has_feb and jan_month:
        jan_data = emp_df[emp_df["Month"] == jan_month].iloc[0]
        feb_data = create_february_row(jan_data, active_numeric_cols)
        
        data = [sr_no, "Feb-2026"]
        for col in active_numeric_cols:
            if col == "PT":
                data.append(feb_data["PT"])
            elif col == "GROUP ACCIDENTAL POLICY":
                data.append(531)
            elif col == "INCOME TAX":
                data.append(0)
            else:
                data.append(get_value(feb_data, col))
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal='right')
        
        sr_no += 1
        current_row += 1
    
    # ===== TOTAL ROW =====
    def safe_sum(col_name):
        try:
            ser = emp_df[col_name].fillna(0)
            if col_name == "INCOME TAX" and "Month" in emp_df.columns:
                ser = ser.where(~emp_df["Month"].isin(feb_variants), 0)
            total = ser.sum()
            # Add February values if applicable
            if has_jan and not has_feb and jan_month:
                jan_data = emp_df[emp_df["Month"] == jan_month].iloc[0]
                feb_data = create_february_row(jan_data, active_numeric_cols)
                if col_name == "PT":
                    total += feb_data["PT"]
                elif col_name == "GROUP ACCIDENTAL POLICY":
                    total += 531
                elif col_name == "INCOME TAX":
                    total += 0
                else:
                    feb_val = get_value(feb_data, col_name)
                    if feb_val and feb_val != "":
                        try:
                            total += float(feb_val)
                        except:
                            pass
            return total
        except:
            return 0
    
    totals = ["", "Total"]
    for col in active_numeric_cols:
        totals.append(safe_sum(col))
    
    for col_idx, value in enumerate(totals, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=value)
        cell.border = thin_border
        cell.font = Font(bold=True)
        if col_idx > 1:
            cell.alignment = Alignment(horizontal='right')
    
    current_row += 2
    
    # ===== FOOTER =====
    ws.cell(row=current_row, column=1, 
            value="या तक्त्यात काही चूक आढळून आल्यास तात्काळ मुख्याध्यापकांच्या लक्षात आणून द्यावी, नजरचुकीने काही चूक झाल्यास लागणा-या आयकरास कर्मचारी स्वत: जबाबदार राहील.")
    current_row += 2
    
    ws.cell(row=current_row, column=1, value="Employee Signature")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    headmaster_col = max(14, total_cols - 3)
    ws.cell(row=current_row, column=headmaster_col, value="Headmaster")
    ws.cell(row=current_row, column=headmaster_col).font = Font(bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=1, value=f"{salutation} {emp_name}")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    
    # ===== COLUMN WIDTHS - Dynamic to fill A4 Landscape =====
    # A4 landscape usable width with margins: ~140 Excel character units
    # We distribute this across all columns
    
    TOTAL_PAGE_WIDTH = 140  # Excel units for A4 landscape with margins
    SR_NO_WIDTH = 6
    MONTH_WIDTH = 11
    
    ws.column_dimensions['A'].width = SR_NO_WIDTH
    ws.column_dimensions['B'].width = MONTH_WIDTH
    
    # Distribute remaining width evenly among data columns
    remaining_cols = total_cols - 2
    if remaining_cols > 0:
        remaining_width = TOTAL_PAGE_WIDTH - SR_NO_WIDTH - MONTH_WIDTH
        col_width = remaining_width / remaining_cols
        # Ensure minimum readable width of 8 and max of 15
        col_width = max(8, min(15, col_width))
        for i in range(3, total_cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = col_width
    
    # ===== PAGE SETUP FOR A4 LANDSCAPE PRINTING =====
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # Fit all content to 1 page
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    
    # Set comfortable margins (in inches)
    ws.page_margins = PageMargins(
        left=0.3, right=0.3,
        top=0.4, bottom=0.4,
        header=0.2, footer=0.2
    )
    
    # Set print area to include ALL content
    last_row = current_row
    last_col = get_column_letter(total_cols)
    ws.print_area = f'A1:{last_col}{last_row}'
    
    # Repeat header row on each page
    ws.print_title_rows = '3:3'
    
    # Center horizontally, align to top (no vertical centering)
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    
    return wb


def create_consolidated_pdf_from_excel(output_dir, output_pdf_path):
    """
    Create consolidated PDF by exporting each Excel file to PDF using Excel,
    then merging all PDFs into one.
    """
    print("\nCreating consolidated PDF using Excel export...")
    
    # Create temp folder for individual PDFs
    temp_pdf_dir = os.path.join(output_dir, "_temp_pdfs")
    os.makedirs(temp_pdf_dir, exist_ok=True)
    
    # Get all xlsx files
    xlsx_files = [f for f in os.listdir(output_dir) if f.endswith('.xlsx')]
    xlsx_files.sort()  # Sort alphabetically by employee name
    
    if not xlsx_files:
        print("No Excel files found to convert!")
        return 0
    
    print(f"Found {len(xlsx_files)} Excel files to convert...")
    
    # Initialize Excel application
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        pdf_files = []
        converted_count = 0
        
        for idx, xlsx_file in enumerate(xlsx_files):
            xlsx_path = os.path.join(output_dir, xlsx_file)
            pdf_name = xlsx_file.replace('.xlsx', '.pdf')
            pdf_path = os.path.join(temp_pdf_dir, pdf_name)
            
            try:
                # Open workbook
                wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
                
                # Export to PDF (0 = xlTypePDF)
                wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
                
                wb.Close(SaveChanges=False)
                
                pdf_files.append(pdf_path)
                converted_count += 1
                
                if converted_count % 10 == 0:
                    print(f"  Converted {converted_count}/{len(xlsx_files)} files...")
                    
            except Exception as e:
                print(f"  Error converting {xlsx_file}: {e}")
                continue
        
        print(f"Converted {converted_count} files to PDF")
        
        # Merge all PDFs
        if pdf_files:
            print("Merging PDFs...")
            merger = PdfMerger()
            
            for pdf_path in pdf_files:
                try:
                    merger.append(pdf_path)
                except Exception as e:
                    print(f"  Error merging {pdf_path}: {e}")
            
            merger.write(output_pdf_path)
            merger.close()
            
            print(f"Merged {len(pdf_files)} PDFs into consolidated file")
        
        # Cleanup temp PDFs
        print("Cleaning up temporary files...")
        for pdf_path in pdf_files:
            try:
                os.remove(pdf_path)
            except:
                pass
        try:
            os.rmdir(temp_pdf_dir)
        except:
            pass
        
        return converted_count
        
    except Exception as e:
        print(f"Excel automation error: {e}")
        raise
    finally:
        if excel:
            try:
                excel.Quit()
            except:
                pass


# =================== MAIN ===================

print("Loading merged data...")
df = load_merged_data()

print(f"Total records loaded: {len(df)}")
print(f"Available columns: {len(df.columns)}")

# Get actual numeric columns that exist in the data (excluding those we don't want)
actual_numeric_cols = [col for col in NUMERIC_COLUMNS if col in df.columns and col not in EXCLUDE_COLUMNS]
print(f"Numeric columns found: {len(actual_numeric_cols)}")

# Find the employee name column
EMP_NAME_COL = "EMPLOYEE NAME"
print(f"Using employee name column: {EMP_NAME_COL}")

# Normalize employee names
df[EMP_NAME_COL] = df[EMP_NAME_COL].astype(str).str.strip().str.upper()

# Remove invalid names
df = df[df[EMP_NAME_COL] != ""]
df = df[df[EMP_NAME_COL] != "NAN"]
df = df[~df[EMP_NAME_COL].str.contains("GRAND TOTAL", case=False, na=False)]

print(f"Unique employees: {df[EMP_NAME_COL].nunique()}")

# Generate reports for each employee
count = 0
for emp_name, emp_df in df.groupby(EMP_NAME_COL):
    if not emp_name or emp_name == "NAN":
        continue
    
    try:
        wb = create_employee_report(emp_name, emp_df, actual_numeric_cols)
        
        # Safe filename
        safe_name = emp_name.replace("/", "_").replace("\\", "_").replace(":", "_")
        safe_name = safe_name.replace("*", "_").replace("?", "_").replace('"', "_")
        safe_name = safe_name.replace("<", "_").replace(">", "_").replace("|", "_")
        
        file_path = os.path.join(OUTPUT_DIR, f"{safe_name}.xlsx")
        wb.save(file_path)
        count += 1
        
        if count % 10 == 0:
            print(f"  Generated {count} reports...")
            
    except Exception as e:
        print(f"Error creating report for {emp_name}: {e}")

print(f"\n[SUCCESS] Generated {count} employee reports in: {OUTPUT_DIR}")

# ===== Generate Consolidated PDF =====
print("\n" + "="*50)
print("Generating Consolidated PDF (Excel Print Preview Style)...")
print("="*50)

pdf_output_path = os.path.join(OUTPUT_DIR, f"{SCHOOL_FOLDER}_All_Reports_Consolidated.pdf")

try:
    pdf_count = create_consolidated_pdf_from_excel(OUTPUT_DIR, pdf_output_path)
    print(f"\n[SUCCESS] Consolidated PDF created: {pdf_output_path}")
    print(f"          Contains {pdf_count} employee reports in landscape A4 format")
except Exception as e:
    print(f"[ERROR] Failed to create consolidated PDF: {e}")
    import traceback
    traceback.print_exc()
