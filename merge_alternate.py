import pandas as pd
import os
import sys
import glob
import xlrd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ===== CONFIGURATION =====
# Usage: python merge_alternate.py [folder_name]
# Example: python merge_alternate.py palasgaon
# Example: python merge_alternate.py "KANYA BASMATH"

ROOT_DIR = os.environ.get("EXCEL_MERGER_ROOT", r"D:\excel merger")

# Get folder name from command line or use default
if len(sys.argv) > 1:
    SCHOOL_FOLDER = sys.argv[1]
else:
    SCHOOL_FOLDER = "palasgaon"  # Default

BASE_PATH = os.path.join(ROOT_DIR, SCHOOL_FOLDER)
OUTPUT_FILE = os.path.join(ROOT_DIR, f"{SCHOOL_FOLDER}_Merged_Monthly.xlsx")

print(f"School: {SCHOOL_FOLDER}")
print(f"Source: {BASE_PATH}")
print(f"Output: {OUTPUT_FILE}")


def find_header_row(file_path):
    preview = pd.read_excel(
        file_path,
        engine="xlrd",
        header=None,
        nrows=40
    )

    for i, row in preview.iterrows():
        if row.astype(str).str.upper().isin(["SR.NO"]).any():
            return i

    raise ValueError(f"SR.NO not found in {file_path}")


def copy_heading_from_source(source_file, dest_ws, header_row):
    """
    Copy heading rows exactly as they are from source .xls file.
    Preserves cell positions, merged cells, and formatting.
    """
    source_wb = xlrd.open_workbook(source_file, formatting_info=True)
    source_sheet = source_wb.sheet_by_index(0)
    
    # Copy cells from heading rows
    for row_idx in range(header_row):
        for col_idx in range(source_sheet.ncols):
            cell = source_sheet.cell(row_idx, col_idx)
            value = cell.value
            
            if value == '' or value is None:
                continue
            
            # Get formatting
            xf_index = cell.xf_index
            xf = source_wb.xf_list[xf_index]
            font = source_wb.font_list[xf.font_index]
            
            # Write to destination (1-indexed)
            dest_cell = dest_ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
            
            # Apply font
            font_size = font.height // 20
            if font_size < 8:
                font_size = 11
            
            dest_cell.font = Font(
                name=font.name if font.name else 'Calibri',
                size=font_size,
                bold=font.bold,
                italic=font.italic
            )
            
            # Alignment
            h_align_map = {0: 'general', 1: 'left', 2: 'center', 3: 'right'}
            h_align = h_align_map.get(xf.alignment.hor_align, 'general')
            dest_cell.alignment = Alignment(horizontal=h_align, vertical='center')
    
    # Copy merged cells
    for (rlo, rhi, clo, chi) in source_sheet.merged_cells:
        if rlo < header_row:
            dest_ws.merge_cells(
                start_row=rlo + 1,
                start_column=clo + 1,
                end_row=min(rhi, header_row),
                end_column=chi
            )
    
    # Copy column widths
    for col_idx in range(source_sheet.ncols):
        col_width = source_sheet.colinfo_map.get(col_idx)
        if col_width:
            width = col_width.width / 256
            dest_ws.column_dimensions[get_column_letter(col_idx + 1)].width = width
    
    # Copy row heights
    for row_idx in range(header_row):
        row_height = source_sheet.rowinfo_map.get(row_idx)
        if row_height:
            dest_ws.row_dimensions[row_idx + 1].height = row_height.height / 20
    
    return header_row


def get_column_header_styles(source_file, header_row):
    """Extract column header styles from source file."""
    workbook = xlrd.open_workbook(source_file, formatting_info=True)
    sheet = workbook.sheet_by_index(0)
    
    styles = []
    for col_idx in range(sheet.ncols):
        cell = sheet.cell(header_row, col_idx)
        xf_index = cell.xf_index
        xf = workbook.xf_list[xf_index]
        font = workbook.font_list[xf.font_index]
        
        # Background color
        pattern_colour_index = xf.background.pattern_colour_index
        bg_color = None
        if pattern_colour_index and pattern_colour_index < len(workbook.colour_map):
            rgb = workbook.colour_map.get(pattern_colour_index)
            if rgb:
                bg_color = '{:02X}{:02X}{:02X}'.format(*rgb)
        
        styles.append({
            'bold': font.bold,
            'italic': font.italic,
            'font_size': font.height // 20,
            'font_name': font.name,
            'bg_color': bg_color,
        })
    
    return styles


def apply_cell_style(cell, style_info=None, is_header=False):
    """Apply style to cell."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border
    
    if is_header and style_info:
        font_size = style_info.get('font_size', 11)
        if font_size < 8:
            font_size = 11
        
        cell.font = Font(
            name=style_info.get('font_name', 'Calibri'),
            size=font_size,
            bold=style_info.get('bold', True),
            italic=style_info.get('italic', False)
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        bg_color = style_info.get('bg_color')
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    elif is_header:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


# =================== MAIN PROCESSING ===================

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# Get heading/styles from first file found
first_file = None
header_row_first = None
column_styles = None

print("Processing monthly sheets...")

for month in sorted(os.listdir(BASE_PATH)):
    month_path = os.path.join(BASE_PATH, month)

    if not os.path.isdir(month_path):
        continue

    files = glob.glob(os.path.join(month_path, "*.xls"))
    if not files:
        continue

    month_tables = []

    for file in files:
        header_row = find_header_row(file)

        # Save first file info
        if first_file is None:
            first_file = file
            header_row_first = header_row
            column_styles = get_column_header_styles(file, header_row)
            print(f"Using heading from: {os.path.basename(file)}")

        df = pd.read_excel(
            file,
            engine="xlrd",
            header=header_row
        )

        df = df.dropna(how="all")

        # Remove Grand Total
        if "SR.NO" in df.columns:
            df = df[df["SR.NO"].astype(str).str.upper() != "GRAND TOTAL"]

        df["Source_File"] = os.path.basename(file)
        month_tables.append(df)

    final_month_df = pd.concat(month_tables, ignore_index=True)

    # Create sheet for this month
    ws = wb.create_sheet(title=month)

    # Copy heading exactly from first source file
    heading_rows = copy_heading_from_source(first_file, ws, header_row_first)

    # Blank row after heading
    current_row = heading_rows + 2

    # Write table data
    for r_idx, r in enumerate(dataframe_to_rows(final_month_df, index=False, header=True)):
        for c_idx, value in enumerate(r, start=1):
            cell = ws.cell(row=current_row, column=c_idx, value=value)
            
            if r_idx == 0:  # Header row
                style = column_styles[c_idx - 1] if c_idx <= len(column_styles) else None
                apply_cell_style(cell, style, is_header=True)
            else:
                apply_cell_style(cell)
        
        current_row += 1

    print(f"  [{month}] - {len(final_month_df)} rows from {len(files)} files")

wb.save(OUTPUT_FILE)

print(f"\n[SUCCESS] Monthly Excel created: {OUTPUT_FILE}")
print(f"   Sheets: {len(wb.sheetnames)}")
