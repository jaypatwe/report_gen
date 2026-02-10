import pandas as pd
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ===== CONFIGURATION =====
# Usage: python summary.py [folder_name]
# Example: python summary.py palasgaon
# Example: python summary.py "KANYA BASMATH"

ROOT_DIR = os.environ.get("EXCEL_MERGER_ROOT", r"D:\excel merger")

if len(sys.argv) > 1:
    SCHOOL_FOLDER = sys.argv[1]
else:
    SCHOOL_FOLDER = "palasgaon"

MERGED_FILE = os.path.join(ROOT_DIR, f"{SCHOOL_FOLDER}_Merged_Monthly.xlsx")
OUTPUT_FILE = os.path.join(ROOT_DIR, f"{SCHOOL_FOLDER}_Summary_Totals.xlsx")

print(f"School: {SCHOOL_FOLDER}")
print(f"Merged File: {MERGED_FILE}")
print(f"Output: {OUTPUT_FILE}")

# Month order for sorting
MONTH_ORDER = [
    "mar 25", "mar25", "MAR 25",
    "apr 25", "apr25", "APR 25",
    "may 25", "may25", "MAY 25",
    "jun 25", "jun25", "JUN 25",
    "jul 25", "jul25", "JUL 25",
    "aug 25", "aug25", "AUG 25",
    "sep 25", "sep25", "SEP 25",
    "oct 25", "oct25", "OCT 25",
    "nov 25", "nov25", "NOV 25",
    "dec 25", "dec25", "DEC 25",
    "jan 26", "jan26", "JAN 26",
    "feb 26", "feb26", "FEB 26"
]

# Columns to exclude from report
EXCLUDE_COLUMNS = [
    "Unnamed: 0", "SR.NO", "Source_File", "Month", "Month_Order",
    "EMPLOYEE NAME", "GENDER M/F", "NAME OF SCHOOL",
    "GROSS AFTER DEDUCTING FA", "GROSS PAYMENT AFTER GOVT DEDUCTIONS",
    "GROSS PAYMENT AFTER NPS DEDUCTIONS", "NGR(TOTAL DEDUCTIONS)",
    "EMPLOYEE NET SALARY", "TOTAL GOVT DEDUCTIONS", "NPS TOTAL"
]

# Numeric columns (same as report.py)
NUMERIC_COLUMNS = [
    "BASIC PAY", "D.A", "HRA", "T.A", "T.A ARREARS", "TRIBAL ALLOWANCE",
    "WASHING ALLOWANCE ", "DA ARREARS ", "HRA ARREARS ", "BASIC ARREARS ",
    "CLA", "NPS EMPR ALLOW", "TOTAL PAY", "F A",
    "GPF", "GPF ADV", "PT", "GIS(ZP)", "GIS SCOUT", "DCPS REGULAR",
    "DCPS DELAYED", "DCPS PAY ARREARS RECOVERY", "REVENUE STAMP",
    "DCPS DA ARREARS RECOVERY", "GROUP ACCIDENTAL POLICY", "NAA",
    "TOTAL GOVT DEDUCTIONS",
    "NPS EMPR CONTRI", "NPS EMP CONTRI", "NPS EMPR CONTRI ARR",
    "NPS EMP CONTRI ARR", "NPS TOTAL",
    "INCOME TAX", "CO-OP BANK", "NGR(LIC)", "NGR(SOCIETY LOAN)", "NGR(MISC)",
    "NGR(OTHER RECOVERY)", "NGR(RD)", "NGR(OTHER DEDUCTION)"
]


# ===== HELPER FUNCTIONS =====

def find_header_row_xlsx(file_path, sheet_name=0):
    """Find the row containing SR.NO header."""
    preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
    for i, row in preview.iterrows():
        if row.astype(str).str.upper().str.strip().isin(["SR.NO"]).any():
            return i
    return 0


def load_merged_data():
    """Load data from the merged Excel file (all sheets)."""
    print(f"Reading {MERGED_FILE}...")

    xlsx = pd.ExcelFile(MERGED_FILE)
    sheet_names = xlsx.sheet_names
    print(f"Found {len(sheet_names)} sheets: {sheet_names}")

    all_data = []

    for sheet_name in sheet_names:
        header_row = find_header_row_xlsx(MERGED_FILE, sheet_name)
        df = pd.read_excel(MERGED_FILE, sheet_name=sheet_name, header=header_row)
        df = df.dropna(how="all")

        if "Month" not in df.columns:
            df["Month"] = sheet_name

        all_data.append(df)
        print(f"  [{sheet_name}] - {len(df)} rows")

    combined_df = pd.concat(all_data, ignore_index=True)
    print(f"\nTotal records: {len(combined_df)}")

    return combined_df


def get_value(row_data, col_name):
    """Safely get value from row."""
    if col_name is None:
        return 0
    try:
        val = row_data.get(col_name, 0)
        if pd.isna(val):
            return 0
        return val
    except:
        return 0


def create_february_row(jan_row, numeric_cols):
    """Create February row based on January data with modifications."""
    feb_row = jan_row.copy()
    feb_row["Month"] = "FEB 26"

    pt_val = get_value(jan_row, "PT")
    if pt_val == 200:
        feb_row["PT"] = 300
    elif pt_val == 0 or pt_val == "" or pd.isna(pt_val):
        feb_row["PT"] = 0
    else:
        feb_row["PT"] = 300

    feb_row["GROUP ACCIDENTAL POLICY"] = 531

    if "INCOME TAX" in feb_row.index:
        feb_row["INCOME TAX"] = 0

    return feb_row


def compute_employee_totals(emp_df, numeric_cols):
    """Compute the total row for one employee (same logic as report.py)."""
    jan_variants = ["jan26", "jan 26", "JAN 26"]
    feb_variants = ["feb26", "feb 26", "FEB 26"]

    emp_df = emp_df.copy()

    # Income Tax must be 0 for February
    if "INCOME TAX" in emp_df.columns and "Month" in emp_df.columns:
        emp_df.loc[emp_df["Month"].isin(feb_variants), "INCOME TAX"] = 0

    has_jan = any(v in emp_df["Month"].values for v in jan_variants)
    has_feb = any(v in emp_df["Month"].values for v in feb_variants)

    jan_month = None
    for v in jan_variants:
        if v in emp_df["Month"].values:
            jan_month = v
            break

    totals = {}
    for col in numeric_cols:
        try:
            ser = emp_df[col].fillna(0) if col in emp_df.columns else pd.Series([0])
            # For INCOME TAX, zero-out February values before summing
            if col == "INCOME TAX" and "Month" in emp_df.columns:
                ser = ser.where(~emp_df["Month"].isin(feb_variants), 0)
            total = ser.sum()

            # Add February values if January exists but February doesn't
            if has_jan and not has_feb and jan_month:
                jan_data = emp_df[emp_df["Month"] == jan_month].iloc[0]
                feb_data = create_february_row(jan_data, numeric_cols)
                if col == "PT":
                    total += feb_data["PT"]
                elif col == "GROUP ACCIDENTAL POLICY":
                    total += 531
                elif col == "INCOME TAX":
                    total += 0
                else:
                    feb_val = get_value(feb_data, col)
                    if feb_val and feb_val != "":
                        try:
                            total += float(feb_val)
                        except:
                            pass
            totals[col] = total
        except:
            totals[col] = 0

    return totals


# =================== MAIN ===================

print("Loading merged data...")
df = load_merged_data()

print(f"Total records loaded: {len(df)}")

# Get actual numeric columns present in the data
actual_numeric_cols = [col for col in NUMERIC_COLUMNS if col in df.columns and col not in EXCLUDE_COLUMNS]
print(f"Numeric columns found: {len(actual_numeric_cols)}")

EMP_NAME_COL = "EMPLOYEE NAME"

# Normalize employee names
df[EMP_NAME_COL] = df[EMP_NAME_COL].astype(str).str.strip().str.upper()
df = df[df[EMP_NAME_COL] != ""]
df = df[df[EMP_NAME_COL] != "NAN"]
df = df[~df[EMP_NAME_COL].str.contains("GRAND TOTAL", case=False, na=False)]

print(f"Unique employees: {df[EMP_NAME_COL].nunique()}")

# Build summary rows
summary_rows = []
for emp_name, emp_df in df.groupby(EMP_NAME_COL):
    if not emp_name or emp_name == "NAN":
        continue

    try:
        totals = compute_employee_totals(emp_df, actual_numeric_cols)
        row = {"EMPLOYEE NAME": emp_name}
        row.update(totals)
        summary_rows.append(row)
    except Exception as e:
        print(f"Error computing totals for {emp_name}: {e}")

print(f"\nCreating summary Excel with {len(summary_rows)} employees...")

# ===== CREATE EXCEL =====
wb = Workbook()
ws = wb.active
ws.title = "Summary Totals"

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(bold=True, size=10)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_text = Font(bold=True, size=10, color="FFFFFF")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Row 1: Title
headers = ["SR.NO", "EMPLOYEE NAME"] + actual_numeric_cols
total_cols = len(headers)

ws.cell(row=1, column=1, value=f"{SCHOOL_FOLDER} - Employee Summary Totals (2025-26)")
ws.cell(row=1, column=1).font = Font(bold=True, size=14)
ws.cell(row=1, column=1).alignment = center_align
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

# Row 2: Column Headers
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.font = header_text
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

# Data rows
for row_idx, row_data in enumerate(summary_rows, start=3):
    sr_no = row_idx - 2

    ws.cell(row=row_idx, column=1, value=sr_no).border = thin_border
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')

    ws.cell(row=row_idx, column=2, value=row_data["EMPLOYEE NAME"]).border = thin_border

    for col_offset, col_name in enumerate(actual_numeric_cols, start=3):
        val = row_data.get(col_name, 0)
        cell = ws.cell(row=row_idx, column=col_offset, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right')

# Grand Total row
grand_total_row = len(summary_rows) + 3
ws.cell(row=grand_total_row, column=1, value="").border = thin_border
ws.cell(row=grand_total_row, column=2, value="GRAND TOTAL").border = thin_border
ws.cell(row=grand_total_row, column=2).font = Font(bold=True)

for col_offset, col_name in enumerate(actual_numeric_cols, start=3):
    col_total = sum(r.get(col_name, 0) for r in summary_rows)
    cell = ws.cell(row=grand_total_row, column=col_offset, value=col_total)
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='right')

# Column widths
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 30
for i in range(3, total_cols + 1):
    ws.column_dimensions[get_column_letter(i)].width = 14

# Freeze top rows and name column so scrolling is easy
ws.freeze_panes = "C3"

# Page setup
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

wb.save(OUTPUT_FILE)
print(f"\n[SUCCESS] Summary file created: {OUTPUT_FILE}")
print(f"   Employees: {len(summary_rows)}")
print(f"   Columns: {len(actual_numeric_cols)}")
