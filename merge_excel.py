import pandas as pd
import os
import glob
import xlrd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

BASE_PATH = r"D:\excel merger\palasgaon"
OUTPUT_FILE = r"D:\excel merger\Final_Merged_Report.xlsx"


def find_header_row(file_path):
    preview = pd.read_excel(
        file_path,
        engine="xlrd",
        header=None,
        nrows=40
    )

    for i, row in preview.iterrows():
        if row.astype(str).str.upper().isin(["SR.NO"]).any():
            return i

    raise ValueError(f"SR.NO header not found in {file_path}")


def copy_heading_from_source(source_file, dest_ws, header_row):
    """
    Copy heading rows exactly as they are from source .xls file to destination worksheet.
    Preserves cell positions, merged cells, and formatting.
    """
    source_wb = xlrd.open_workbook(source_file, formatting_info=True)
    source_sheet = source_wb.sheet_by_index(0)
    
    # Copy cells from heading rows (before header_row)
    for row_idx in range(header_row):
        for col_idx in range(source_sheet.ncols):
            cell = source_sheet.cell(row_idx, col_idx)
            value = cell.value
            
            # Skip empty cells
            if value == '' or value is None:
                continue
            
            # Get cell formatting
            xf_index = cell.xf_index
            xf = source_wb.xf_list[xf_index]
            font = source_wb.font_list[xf.font_index]
            
            # Write to destination (openpyxl is 1-indexed)
            dest_cell = dest_ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
            
            # Apply font style
            font_size = font.height // 20
            if font_size < 8:
                font_size = 11
            
            dest_cell.font = Font(
                name=font.name if font.name else 'Calibri',
                size=font_size,
                bold=font.bold,
                italic=font.italic
            )
            
            # Apply alignment
            h_align_map = {0: 'general', 1: 'left', 2: 'center', 3: 'right'}
            h_align = h_align_map.get(xf.alignment.hor_align, 'general')
            dest_cell.alignment = Alignment(horizontal=h_align, vertical='center')
    
    # Copy merged cells from heading area
    for (rlo, rhi, clo, chi) in source_sheet.merged_cells:
        # Only copy merges that are in the heading area
        if rlo < header_row:
            # openpyxl uses 1-based indexing
            dest_ws.merge_cells(
                start_row=rlo + 1,
                start_column=clo + 1,
                end_row=min(rhi, header_row),  # Don't merge past header
                end_column=chi
            )
    
    # Copy column widths
    for col_idx in range(source_sheet.ncols):
        col_width = source_sheet.colinfo_map.get(col_idx)
        if col_width:
            # Convert xlrd width to openpyxl width (approximate)
            width = col_width.width / 256
            dest_ws.column_dimensions[get_column_letter(col_idx + 1)].width = width
    
    # Copy row heights for heading rows
    for row_idx in range(header_row):
        row_height = source_sheet.rowinfo_map.get(row_idx)
        if row_height:
            dest_ws.row_dimensions[row_idx + 1].height = row_height.height / 20
    
    return header_row


def copy_column_header_style(source_file, header_row):
    """Extract column header styles from source file."""
    workbook = xlrd.open_workbook(source_file, formatting_info=True)
    sheet = workbook.sheet_by_index(0)
    
    header_styles = []
    
    for col_idx in range(sheet.ncols):
        cell = sheet.cell(header_row, col_idx)
        xf_index = cell.xf_index
        xf = workbook.xf_list[xf_index]
        font = workbook.font_list[xf.font_index]
        
        # Get background color
        pattern_colour_index = xf.background.pattern_colour_index
        bg_color = None
        if pattern_colour_index and pattern_colour_index < len(workbook.colour_map):
            rgb = workbook.colour_map.get(pattern_colour_index)
            if rgb:
                bg_color = '{:02X}{:02X}{:02X}'.format(*rgb)
        
        style_info = {
            'bold': font.bold,
            'italic': font.italic,
            'font_size': font.height // 20,
            'font_name': font.name,
            'bg_color': bg_color,
        }
        header_styles.append(style_info)
    
    return header_styles


def apply_header_style(cell, style_info):
    """Apply extracted style to openpyxl cell for column headers."""
    font_name = style_info.get('font_name', 'Calibri')
    font_size = style_info.get('font_size', 11)
    if font_size < 8:
        font_size = 11
    
    cell.font = Font(
        name=font_name,
        size=font_size,
        bold=style_info.get('bold', True),
        italic=style_info.get('italic', False)
    )
    
    # Border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Background color
    bg_color = style_info.get('bg_color')
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')


# =================== MAIN PROCESSING ===================

all_tables = []
first_file = None
header_row_first = None
column_header_styles = None

print("Scanning files...")

for month in sorted(os.listdir(BASE_PATH)):
    month_path = os.path.join(BASE_PATH, month)

    if not os.path.isdir(month_path):
        continue

    files = glob.glob(os.path.join(month_path, "*.xls"))

    for file in files:
        header_row = find_header_row(file)

        # Save first file info for copying heading
        if first_file is None:
            first_file = file
            header_row_first = header_row
            column_header_styles = copy_column_header_style(file, header_row)
            print(f"Using heading from: {os.path.basename(file)}")

        df = pd.read_excel(
            file,
            engine="xlrd",
            header=header_row
        )

        df = df.dropna(how="all")

        # Remove Grand Total rows
        if "SR.NO" in df.columns:
            df = df[df["SR.NO"].astype(str).str.upper() != "GRAND TOTAL"]

        # Metadata
        df["Month"] = month
        df["Source_File"] = os.path.basename(file)

        all_tables.append(df)
        print(f"  Processed: {month}/{os.path.basename(file)}")

final_df = pd.concat(all_tables, ignore_index=True)

# ================= WRITE FINAL EXCEL WITH EXACT HEADING =================

print("\nCreating output with original heading...")

wb = Workbook()
ws = wb.active
ws.title = "Consolidated Report"

# Copy heading exactly from source file
heading_rows = copy_heading_from_source(first_file, ws, header_row_first)

# Leave a blank row after heading
current_row = heading_rows + 2

# Track where table header starts
table_header_row = current_row

# Write table data
for r_idx, r in enumerate(dataframe_to_rows(final_df, index=False, header=True)):
    for c_idx, value in enumerate(r, start=1):
        cell = ws.cell(row=current_row, column=c_idx, value=value)
        
        if r_idx == 0:  # This is the header row
            # Apply column header style
            if c_idx <= len(column_header_styles):
                apply_header_style(cell, column_header_styles[c_idx - 1])
            else:
                # For new columns (Month, Source_File), apply default header style
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        else:
            # Apply thin border to data cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    current_row += 1

# Adjust column widths for data columns (keep heading column widths)
for col_idx, column in enumerate(final_df.columns, start=1):
    max_length = len(str(column))
    for row in range(table_header_row + 1, current_row):
        cell_value = ws.cell(row=row, column=col_idx).value
        if cell_value:
            max_length = max(max_length, len(str(cell_value)))
    
    # Only adjust if wider than current
    current_width = ws.column_dimensions[get_column_letter(col_idx)].width or 8
    new_width = min(max_length + 2, 50)
    if new_width > current_width:
        ws.column_dimensions[get_column_letter(col_idx)].width = new_width

wb.save(OUTPUT_FILE)

print(f"\n[SUCCESS] Final merged Excel created: {OUTPUT_FILE}")
print(f"   Total rows: {len(final_df)}")
print(f"   Total columns: {len(final_df.columns)}")
